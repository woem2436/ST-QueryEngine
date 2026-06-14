import json
import sqlite3
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import openpyxl

from .text_utils import (
    append_unit_if_needed,
    dedupe_keep_order,
    display_value,
    extract_numbers,
    extract_unit,
    normalize_text,
    token_counter,
)


@dataclass
class CellRecord:
    table_id: int
    sheet_name: str
    row: int
    col: int
    value: str
    raw_value: Any
    row_text: str
    above_text: str
    left_text: str
    right_text: str
    title: str
    unit: str = ""
    tokens: Dict[str, int] = field(default_factory=dict)

    @property
    def context(self) -> str:
        return " ".join(
            part
            for part in [
                self.title,
                self.above_text,
                self.left_text,
                self.right_text,
                self.row_text,
                self.value,
            ]
            if part
        )

    @property
    def answer(self) -> str:
        return append_unit_if_needed(self.value, self.unit)

    @property
    def is_numeric_like(self) -> bool:
        return bool(extract_numbers(self.value)) or any(mark in self.value for mark in "%≥≧≤<>")

    def to_json(self) -> Dict[str, Any]:
        return {
            "table_id": self.table_id,
            "sheet_name": self.sheet_name,
            "row": self.row,
            "col": self.col,
            "value": self.value,
            "raw_value": _json_safe(self.raw_value),
            "row_text": self.row_text,
            "above_text": self.above_text,
            "left_text": self.left_text,
            "right_text": self.right_text,
            "title": self.title,
            "unit": self.unit,
            "tokens": self.tokens,
        }

    @classmethod
    def from_json(cls, data: Dict[str, Any]) -> "CellRecord":
        return cls(**data)


@dataclass
class RowRecord:
    table_id: int
    sheet_name: str
    row: int
    values: List[str]
    title: str
    tokens: Dict[str, int] = field(default_factory=dict)

    @property
    def text(self) -> str:
        return " | ".join([self.title] + [value for value in self.values if value])

    def to_json(self) -> Dict[str, Any]:
        return {
            "table_id": self.table_id,
            "sheet_name": self.sheet_name,
            "row": self.row,
            "values": self.values,
            "title": self.title,
            "tokens": self.tokens,
        }

    @classmethod
    def from_json(cls, data: Dict[str, Any]) -> "RowRecord":
        return cls(**data)


class TableIndexer:
    """Builds a reproducible SQLite + JSON evidence index for semi-structured Excel tables."""

    def __init__(self, raw_dir: str, sqlite_path: str, index_path: str):
        self.raw_dir = Path(raw_dir)
        self.sqlite_path = Path(sqlite_path)
        self.index_path = Path(index_path)
        self.sqlite_path.parent.mkdir(parents=True, exist_ok=True)
        self.index_path.parent.mkdir(parents=True, exist_ok=True)

    def ensure_index(self, rebuild: bool = False) -> Dict[str, Any]:
        if rebuild or not self.index_path.exists() or not self.sqlite_path.exists():
            return self.build()
        return self.load()

    def load(self) -> Dict[str, Any]:
        with open(self.index_path, "r", encoding="utf-8") as handle:
            payload = json.load(handle)
        payload["cells"] = [CellRecord.from_json(item) for item in payload.get("cells", [])]
        payload["rows"] = [RowRecord.from_json(item) for item in payload.get("rows", [])]
        return payload

    def build(self) -> Dict[str, Any]:
        tables: Dict[str, Dict[str, Any]] = {}
        cells: List[CellRecord] = []
        rows: List[RowRecord] = []

        for excel_file in self._iter_excel_files():
            table_id = int(excel_file.stem)
            table_cells, table_rows, metadata = self._parse_workbook(excel_file, table_id)
            cells.extend(table_cells)
            rows.extend(table_rows)
            tables[str(table_id)] = metadata

        payload = {
            "version": 2,
            "tables": tables,
            "cells": cells,
            "rows": rows,
        }
        self._write_json(payload)
        self._write_sqlite(tables, cells, rows)
        return payload

    def _iter_excel_files(self) -> Iterable[Path]:
        files = []
        for path in self.raw_dir.glob("*.xlsx"):
            if path.stem.isdigit():
                files.append(path)
        return sorted(files, key=lambda item: int(item.stem))

    def _parse_workbook(self, excel_file: Path, table_id: int):
        workbook = openpyxl.load_workbook(excel_file, data_only=True)
        table_cells: List[CellRecord] = []
        table_rows: List[RowRecord] = []
        titles: List[str] = []
        sheet_meta: List[Dict[str, Any]] = []

        for sheet in workbook.worksheets:
            matrix, raw_matrix = self._sheet_matrix(sheet)
            matrix, raw_matrix = self._trim_empty_edges(matrix, raw_matrix)
            if not matrix:
                continue
            title = self._detect_title(matrix)
            if title:
                titles.append(title)
            sheet_meta.append(
                {
                    "sheet_name": sheet.title,
                    "rows": len(matrix),
                    "columns": len(matrix[0]) if matrix else 0,
                    "title": title,
                }
            )

            for row_idx, row_values in enumerate(matrix, start=1):
                compact_values = [value for value in row_values if value]
                if not compact_values:
                    continue
                row_record = RowRecord(
                    table_id=table_id,
                    sheet_name=sheet.title,
                    row=row_idx,
                    values=row_values,
                    title=title,
                )
                row_record.tokens = dict(token_counter(row_record.text))
                table_rows.append(row_record)

                row_text = " | ".join(compact_values)
                for col_idx, value in enumerate(row_values, start=1):
                    if not value:
                        continue
                    above_values = self._above_values(matrix, row_idx - 1, col_idx - 1)
                    left_values = [item for item in row_values[: col_idx - 1] if item]
                    right_values = [item for item in row_values[col_idx:] if item]
                    context_for_unit = " ".join(above_values + left_values + [row_text])
                    record = CellRecord(
                        table_id=table_id,
                        sheet_name=sheet.title,
                        row=row_idx,
                        col=col_idx,
                        value=value,
                        raw_value=raw_matrix[row_idx - 1][col_idx - 1],
                        row_text=row_text,
                        above_text=" | ".join(above_values),
                        left_text=" | ".join(left_values),
                        right_text=" | ".join(right_values),
                        title=title,
                        unit=extract_unit(context_for_unit),
                    )
                    record.tokens = dict(token_counter(record.context))
                    table_cells.append(record)

        metadata = {
            "file": str(excel_file),
            "table_id": table_id,
            "title": dedupe_keep_order(titles)[0] if titles else excel_file.stem,
            "sheets": sheet_meta,
            "cell_count": len(table_cells),
            "row_count": len(table_rows),
        }
        return table_cells, table_rows, metadata

    def _sheet_matrix(self, sheet):
        matrix: List[List[str]] = []
        raw_matrix: List[List[Any]] = []
        for row_idx in range(1, sheet.max_row + 1):
            row_display: List[str] = []
            row_raw: List[Any] = []
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                row_display.append(display_value(cell.value, cell.number_format))
                row_raw.append(cell.value)
            matrix.append(row_display)
            raw_matrix.append(row_raw)

        for merged_range in sheet.merged_cells.ranges:
            top_left = sheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            display = display_value(top_left.value, top_left.number_format)
            raw = top_left.value
            for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
                for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                    matrix[row_idx - 1][col_idx - 1] = display
                    raw_matrix[row_idx - 1][col_idx - 1] = raw
        return matrix, raw_matrix

    def _trim_empty_edges(self, matrix: List[List[str]], raw_matrix: List[List[Any]]):
        if not matrix:
            return matrix, raw_matrix
        non_empty_rows = [idx for idx, row in enumerate(matrix) if any(row)]
        if not non_empty_rows:
            return [], []
        non_empty_cols = [
            idx
            for idx in range(len(matrix[0]))
            if any(row[idx] for row in matrix if idx < len(row))
        ]
        row_start, row_end = min(non_empty_rows), max(non_empty_rows)
        col_start, col_end = min(non_empty_cols), max(non_empty_cols)
        trimmed = [row[col_start : col_end + 1] for row in matrix[row_start : row_end + 1]]
        trimmed_raw = [row[col_start : col_end + 1] for row in raw_matrix[row_start : row_end + 1]]
        return trimmed, trimmed_raw

    def _detect_title(self, matrix: List[List[str]]) -> str:
        for row in matrix[:5]:
            values = dedupe_keep_order([value for value in row if value])
            if not values:
                continue
            if len(values) == 1:
                return values[0]
            return " ".join(values[:3])
        return ""

    def _above_values(self, matrix: List[List[str]], row_idx: int, col_idx: int) -> List[str]:
        values: List[str] = []
        for idx in range(0, min(20, row_idx)):
            value = matrix[idx][col_idx]
            if value:
                values.append(value)
        for idx in range(max(0, row_idx - 10), row_idx):
            value = matrix[idx][col_idx]
            if value:
                values.append(value)
        return dedupe_keep_order(values)

    def _write_json(self, payload: Dict[str, Any]) -> None:
        serializable = {
            "version": payload["version"],
            "tables": payload["tables"],
            "cells": [cell.to_json() for cell in payload["cells"]],
            "rows": [row.to_json() for row in payload["rows"]],
        }
        with open(self.index_path, "w", encoding="utf-8") as handle:
            json.dump(serializable, handle, ensure_ascii=False, indent=2)

    def _write_sqlite(
        self,
        tables: Dict[str, Dict[str, Any]],
        cells: List[CellRecord],
        rows: List[RowRecord],
    ) -> None:
        conn = sqlite3.connect(str(self.sqlite_path))
        cur = conn.cursor()
        cur.executescript(
            """
            DROP TABLE IF EXISTS table_metadata;
            DROP TABLE IF EXISTS cell_records;
            DROP TABLE IF EXISTS row_records;

            CREATE TABLE table_metadata (
                table_id INTEGER PRIMARY KEY,
                title TEXT,
                file_path TEXT,
                row_count INTEGER,
                cell_count INTEGER
            );

            CREATE TABLE cell_records (
                table_id INTEGER,
                sheet_name TEXT,
                row_index INTEGER,
                col_index INTEGER,
                value TEXT,
                answer TEXT,
                unit TEXT,
                row_text TEXT,
                above_text TEXT,
                left_text TEXT,
                right_text TEXT,
                numeric_value REAL
            );

            CREATE TABLE row_records (
                table_id INTEGER,
                sheet_name TEXT,
                row_index INTEGER,
                row_text TEXT
            );
            """
        )
        cur.executemany(
            """
            INSERT INTO table_metadata(table_id, title, file_path, row_count, cell_count)
            VALUES (?, ?, ?, ?, ?)
            """,
            [
                (
                    int(table_id),
                    meta.get("title", ""),
                    meta.get("file", ""),
                    meta.get("row_count", 0),
                    meta.get("cell_count", 0),
                )
                for table_id, meta in tables.items()
            ],
        )
        cur.executemany(
            """
            INSERT INTO cell_records(
                table_id, sheet_name, row_index, col_index, value, answer, unit,
                row_text, above_text, left_text, right_text, numeric_value
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    cell.table_id,
                    cell.sheet_name,
                    cell.row,
                    cell.col,
                    cell.value,
                    cell.answer,
                    cell.unit,
                    cell.row_text,
                    cell.above_text,
                    cell.left_text,
                    cell.right_text,
                    extract_numbers(cell.value)[0] if extract_numbers(cell.value) else None,
                )
                for cell in cells
            ],
        )
        cur.executemany(
            """
            INSERT INTO row_records(table_id, sheet_name, row_index, row_text)
            VALUES (?, ?, ?, ?)
            """,
            [(row.table_id, row.sheet_name, row.row, row.text) for row in rows],
        )
        conn.commit()
        conn.close()


def _json_safe(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value
