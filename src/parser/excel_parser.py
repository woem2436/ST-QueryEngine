from pathlib import Path
from typing import Any, List, Optional, Tuple

import openpyxl
import pandas as pd

from src.text_utils import display_value


class ExcelParser:
    """半结构化表格的 Excel 解析器。

    它会填充合并单元格、移除外围全空行列，并返回保留原始视觉结构的
    DataFrame，避免过早把所有 sheet 强行压平成单一关系表。
    """

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        self.wb = openpyxl.load_workbook(self.file_path, data_only=True)

    def parse_sheet(self, sheet_name: Optional[str] = None) -> pd.DataFrame:
        matrix, _ = self.parse_sheet_matrix(sheet_name)
        if not matrix:
            return pd.DataFrame()
        df = pd.DataFrame(matrix)
        header_row = self._detect_header_row(df)
        if header_row is not None:
            headers = [str(value) if value else f"col_{idx + 1}" for idx, value in enumerate(df.iloc[header_row])]
            data = df.iloc[header_row + 1 :].reset_index(drop=True)
            data.columns = headers
            return data
        return df

    def parse_sheet_matrix(self, sheet_name: Optional[str] = None) -> Tuple[List[List[str]], List[List[Any]]]:
        ws = self.wb[sheet_name] if sheet_name else self.wb.active
        display_matrix: List[List[str]] = []
        raw_matrix: List[List[Any]] = []
        for row_idx in range(1, ws.max_row + 1):
            display_row: List[str] = []
            raw_row: List[Any] = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                display_row.append(display_value(cell.value, cell.number_format))
                raw_row.append(cell.value)
            display_matrix.append(display_row)
            raw_matrix.append(raw_row)

        self._fill_merged_cells(ws, display_matrix, raw_matrix)
        return self._trim_empty_edges(display_matrix, raw_matrix)

    def _fill_merged_cells(self, ws, display_matrix, raw_matrix):
        for merged_range in ws.merged_cells.ranges:
            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            display = display_value(top_left_cell.value, top_left_cell.number_format)
            raw_value = top_left_cell.value
            for row_idx in range(merged_range.min_row, merged_range.max_row + 1):
                for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                    display_matrix[row_idx - 1][col_idx - 1] = display
                    raw_matrix[row_idx - 1][col_idx - 1] = raw_value

    def _trim_empty_edges(self, display_matrix, raw_matrix):
        if not display_matrix:
            return [], []
        non_empty_rows = [idx for idx, row in enumerate(display_matrix) if any(row)]
        if not non_empty_rows:
            return [], []
        non_empty_cols = [
            idx
            for idx in range(len(display_matrix[0]))
            if any(row[idx] for row in display_matrix if idx < len(row))
        ]
        row_start, row_end = min(non_empty_rows), max(non_empty_rows)
        col_start, col_end = min(non_empty_cols), max(non_empty_cols)
        display_trimmed = [row[col_start : col_end + 1] for row in display_matrix[row_start : row_end + 1]]
        raw_trimmed = [row[col_start : col_end + 1] for row in raw_matrix[row_start : row_end + 1]]
        return display_trimmed, raw_trimmed

    def _detect_header_row(self, df: pd.DataFrame) -> Optional[int]:
        best_idx = None
        best_score = 0
        for idx in range(min(len(df), 8)):
            row = df.iloc[idx]
            values = [value for value in row if value not in (None, "")]
            if not values:
                continue
            text_count = sum(1 for value in values if not _is_number_like(value))
            unique_count = len(set(map(str, values)))
            score = text_count + unique_count
            if score > best_score and text_count >= max(1, len(values) // 2):
                best_idx = idx
                best_score = score
        return best_idx

    def get_all_sheet_names(self) -> List[str]:
        return self.wb.sheetnames


def _is_number_like(value: Any) -> bool:
    try:
        float(str(value).replace(",", "").replace("%", ""))
        return True
    except ValueError:
        return False
